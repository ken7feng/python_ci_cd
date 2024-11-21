import os
from tkinter import Tk, filedialog, Label, Entry, Button, Spinbox, messagebox
from openpyxl import load_workbook

def upload_file():
    global file_path
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if file_path:
        file_label.config(text=f"已上传文件: {os.path.basename(file_path)}")
        read_columns()

def read_columns():
    global column_headers
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 提取用户选择的头部行号
        selected_row = int(row_spinbox.get())
        column_headers = [cell.value for cell in sheet[selected_row]]

        update_inputs(column_headers)
    except Exception as e:
        messagebox.showerror("错误", f"无法读取文件: {e}")

def update_inputs(headers):
    for widget in input_frame.winfo_children():
        widget.destroy()
    global input_entries
    input_entries = []
    Label(input_frame, text="请输入新的标题:").grid(row=0, column=0, columnspan=2)
    for idx, header in enumerate(headers):
        Label(input_frame, text=f"列 {idx+1} 原标题: {header}").grid(row=idx+1, column=0, sticky="w")
        entry = Entry(input_frame, width=30)
        entry.grid(row=idx+1, column=1)
        input_entries.append(entry)
    Button(input_frame, text="应用修改", command=apply_changes).grid(row=len(headers)+1, column=0, columnspan=2, pady=10)

def apply_changes():
    new_headers = [entry.get() for entry in input_entries]
    if not all(new_headers):
        messagebox.showwarning("警告", "所有列标题都必须填写！")
        return
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 数据整体下移一行
        sheet.insert_rows(1)

        # 应用新的列标题
        for col_num, new_header in enumerate(new_headers, start=1):
            sheet.cell(row=1, column=col_num, value=new_header)

        # 锁定第一行标题
        sheet.freeze_panes = "A2"

        # 另存文件
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if save_path:
            workbook.save(save_path)
            messagebox.showinfo("成功", f"文件已保存到: {save_path}")
    except Exception as e:
        messagebox.showerror("错误", f"应用修改时出错: {e}")

# GUI 构建
root = Tk()
root.title("Excel 列标题修改工具")

file_path = ""
column_headers = []
input_entries = []

# 文件选择部分
file_frame = Label(root)
file_frame.pack(pady=10)

file_label = Label(file_frame, text="请选择一个 Excel 文件...")
file_label.pack()

file_button = Button(file_frame, text="上传文件", command=upload_file)
file_button.pack()

# 行号选择部分
row_frame = Label(root)
row_frame.pack(pady=10)

Label(row_frame, text="请选择头部行号:").pack(side="left")
row_spinbox = Spinbox(row_frame, from_=1, to=100, width=5)  # 假定表格不超过 100 行
row_spinbox.pack(side="left")
row_spinbox.delete(0, "end")
row_spinbox.insert(0, "1")  # 默认使用第一行作为头部

# 列标题输入部分
input_frame = Label(root)
input_frame.pack(pady=10)

# 启动 GUI
root.geometry("500x500")
root.mainloop()
